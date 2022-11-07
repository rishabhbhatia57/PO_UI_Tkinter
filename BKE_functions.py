import os
import shutil
from datetime import datetime
import pandas as pd
import xlsxwriter
import glob
import numpy as np
from openpyxl import load_workbook, Workbook
import openpyxl.utils.cell
import time
from config import ConfigFolderPath, CLIENTSFOLDERPATH
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import Alignment, Font
from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder
from openpyxl.utils import get_column_letter
# from openpyxl.styles import Style
import tabula
import csv
import json
import sys
from openpyxl.styles import PatternFill
from flask import jsonify

import BKE_log
from config import ITEMMASTERPATH, IGSTMASTERPATH, SGSTMASTERPATH, LOCATIONMASTERPATH, LOCATION2MASTERPATH, CLOSINGSTOCKMASTERPATH, get_client_code, get_client_name, REQSUMTEMPLATEPATH, TEMPLATESPATH, PACKINGSLIPTEMPLATEPATH
pd.options.mode.chained_assignment = None
# import warnings
# warnings.simplefilter(action='ignore', category=SettingWithCopyWarning)

logger = BKE_log.setup_custom_logger('root')


def validate_column_names(df_formula_cols, df_master, file_name):
    try:
        print('Validating '+str(file_name)+'...')
        logger.info('Validating '+str(file_name)+'...')
        df_master_columns = pd.DataFrame(
            list(df_master.columns), columns=['Column Names'])
        df_validate_columns = df_formula_cols[df_formula_cols['Master Files'] == file_name]
        df_temp = df_validate_columns.assign(Check=df_validate_columns['Column Names'].isin
                                             (df_master_columns['Column Names']).astype(str))

        df_temp = df_temp[df_temp['Check'].str.contains('False')]

        if not df_temp.empty:
            cols_not_found_list = list(df_temp['Column Names'])
            results = {
                "cols_name": cols_not_found_list,
                "cols_not_found": True
            }
            logger.error('Process Stopped! Unable to find '+str(results['cols_name'])+' columns in '+str(
                file_name)+' file. Kindly check '+str(file_name)+'.')
            return results
        else:

            results = {
                "cols_name": "",
                "cols_not_found": False
            }
            return results
    except Exception as e:
        print(e)
        exc_type, exc_obj, exc_tb = sys.exc_info()
        fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
        print(exc_type, fname, exc_tb.tb_lineno)


def po_check_master_files(formulaWorksheet):

    valid_result = {
        'valid': True
    }

    invalid_result = {
        'valid': False
    }
    try:
        logger.info('Validating master files...')
        df_formula_cols = pd.read_excel(
            formulaWorksheet, sheet_name="Validate Column Names")

        with open(ConfigFolderPath, 'r') as jsonFile:
            config = json.load(jsonFile)

            file_name = "Item Master"
            df_item_master = pd.read_excel(
                ITEMMASTERPATH, sheet_name='Item Master')
            # checking  Item Master 1
            result = validate_column_names(
                df_formula_cols, df_item_master, file_name)
            if result['cols_not_found'] == True:  # Problem exists
                return invalid_result

             # checking  Location Master 2
            file_name = "Location Master"
            df_location_master = pd.read_excel(LOCATIONMASTERPATH, sheet_name='Location Master')
            result1 = validate_column_names(
                df_formula_cols, df_location_master, file_name)
            if result1['cols_not_found'] == True:
                return invalid_result

            # checking  Location Master 3
            file_name = "WH Closing Stock"
            df_closing_stock = pd.read_excel(CLOSINGSTOCKMASTERPATH, sheet_name='ClosingStock', skiprows=3)  # ClosingStock
            result2 = validate_column_names(
                df_formula_cols, df_closing_stock, file_name)
            if result2['cols_not_found'] == True:
                return invalid_result

            # # checking  IGST Master 4
            # file_name = "IGST Master"
            # df_igst_master = pd.read_excel(
            #     config['IGSTMASTERPATH'], sheet_name='DBF')
            # result3 = validate_column_names(
            #     df_formula_cols, df_igst_master, file_name)
            # if result3['cols_not_found'] == True:
            #     return invalid_result

            # # checking  SGST Master 5
            # file_name = "SGST Master"
            # df_sgst_master = pd.read_excel(
            #     config['SGSTMASTERPATH'], sheet_name='DBF')
            # result4 = validate_column_names(
            #     df_formula_cols, df_sgst_master, file_name)
            # if result4['cols_not_found'] == True:
            #     return invalid_result

            # checking  Location 2 Master 6
            file_name = "Location 2 Master"
            df_location_2_master = pd.read_excel(LOCATION2MASTERPATH, sheet_name='Location2')
            result5 = validate_column_names(
                df_formula_cols, df_location_2_master, file_name)
            if result5['cols_not_found'] == True:
                return invalid_result

        return valid_result

    except Exception as e:
        print(e)
        exc_type, exc_obj, exc_tb = sys.exc_info()
        fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
        print(exc_type, fname, exc_tb.tb_lineno)
        logger.error("Error while checking "+str(file_name)+" file: "+str(e))
        return invalid_result


def pkg_check_master_files(formulaWorksheet):

    valid_result = {
        'valid': True
    }

    invalid_result = {
        'valid': False
    }
    try:
        logger.info('Validating master files...')
        df_formula_cols = pd.read_excel(
            formulaWorksheet, sheet_name="Validate Column Names")

        with open(ConfigFolderPath, 'r') as jsonFile:
            config = json.load(jsonFile)

            file_name = "Item Master"
            df_item_master = pd.read_excel(
                ITEMMASTERPATH, sheet_name='Item Master')
            # checking  Item Master 1
            result = validate_column_names(
                df_formula_cols, df_item_master, file_name)
            if result['cols_not_found'] == True:  # Problem exists
                return invalid_result

            #  # checking  Location Master 2
            # file_name = "Location Master"
            # df_location_master = pd.read_excel(LOCATION2MASTERPATH, sheet_name='Location Master')
            # result1 = validate_column_names(
            #     df_formula_cols, df_location_master, file_name)
            # if result1['cols_not_found'] == True:
            #     return invalid_result

            # # checking  Location Master 3
            # file_name = "WH Closing Stock"
            # df_closing_stock = pd.read_excel(CLOSINGSTOCKMASTERPATH, sheet_name='ClosingStock', skiprows=3)  # ClosingStock
            # result2 = validate_column_names(
            #     df_formula_cols, df_closing_stock, file_name)
            # if result2['cols_not_found'] == True:
            #     return invalid_result

            # checking  IGST Master 4
            file_name = "IGST Master"
            df_igst_master = pd.read_excel( IGSTMASTERPATH , sheet_name='DBF')
            result3 = validate_column_names(
                df_formula_cols, df_igst_master, file_name)
            if result3['cols_not_found'] == True:
                return invalid_result

            # checking  SGST Master 5
            file_name = "SGST Master"
            df_sgst_master = pd.read_excel(SGSTMASTERPATH, sheet_name='DBF')
            result4 = validate_column_names(
                df_formula_cols, df_sgst_master, file_name)
            if result4['cols_not_found'] == True:
                return invalid_result

            # checking  Location 2 Master 6
            file_name = "Location 2 Master"
            df_location_2_master = pd.read_excel(LOCATION2MASTERPATH, sheet_name='Location2')
            result5 = validate_column_names(
                df_formula_cols, df_location_2_master, file_name)
            if result5['cols_not_found'] == True:
                return invalid_result

        return valid_result

    except Exception as e:
        print(e)
        exc_type, exc_obj, exc_tb = sys.exc_info()
        fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
        print(exc_type, fname, exc_tb.tb_lineno)
        logger.error("Error while checking "+str(file_name)+" file: "+str(e))
        return invalid_result


def downloadFiles(RootFolder, POSource, OrderDate, ClientCode, base_path):

    source_folder = POSource
    destination_folder = base_path + "/99-Working/10-Download-Files/"
    try:
        logger.info("Copying PDF Files from '"+str(source_folder) +
                    "' to '"+str(destination_folder)+"'")

        for file_name in os.listdir(POSource):
            # construct full file path

            source = source_folder + "/" + file_name
            destination = destination_folder
            # copy only files
            if os.path.isfile(source) and file_name.endswith('.pdf'):
                shutil.copy(source, destination)
                # from source '"+source_folder+"' to destination '"+destination_folder+"'")
                logger.info("File '"+file_name+"' copied")
                # from source '"+source_folder+"' to destination '"+destination_folder+"'")
                print("File '"+file_name+"' copied")
    except Exception as e:
        logger.error("Error while copying files: "+str(e))
        exc_type, exc_obj, exc_tb = sys.exc_info()
        fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
        print(exc_type, fname, exc_tb.tb_lineno)
        print("Error while copying files: "+str(e))


def scriptStarted():
    logger.info('Starting Process...')
    print('Starting Process...')
    print('Script is running...')
    print('Do not close this window while processing...')
    logger.info('Do not close this window while processing...')
    return "Script Started."


def scriptEnded():
    logger.info('Process Ended.')
    print('Process Ended.')
    return "Script Ended"


def checkFolderStructure(RootFolder, ClientCode, OrderDate, mode, base_path):

    try:
        DatedPath = base_path
        isExist = os.path.exists(DatedPath)
        working_99_Exists = os.path.exists(DatedPath+'/99-Working')

        internalDir = ["/99-Working/10-Download-Files", "/99-Working/20-Intermediate-Files",
                       "/99-Working/30-Extract-CSV", "/99-Working/40-Extract-Excel",
                       "/50-Consolidate-Orders", "/60-Requirement-Summary", "/70-Packing-Slip"]

        if mode == 'consolidation':
            if not isExist:
                logger.info("Creating the new directory...")
                os.makedirs(DatedPath)
                for i in range(len(internalDir)):
                    if not os.path.exists(DatedPath+internalDir[i]):
                        os.makedirs(DatedPath+internalDir[i])

            if isExist:
                for i in range(len(internalDir)):
                    if not os.path.exists(DatedPath+internalDir[i]):
                        os.makedirs(DatedPath+internalDir[i])

        if mode == 'packing':
            if not isExist:
                logger.info("Creating the new directory...")
                os.makedirs(DatedPath)
                for i in range(6, len(internalDir)):
                    if not os.path.exists(DatedPath+internalDir[i]):
                        os.makedirs(DatedPath+internalDir[i])
                # os.makedirs(DatedPath+"/70-Packing-Slip")
            if isExist:
                logger.info("Creating the new directory...")
                print("Creating the new directory...")
                for i in range(6, len(internalDir)):
                    if not os.path.exists(DatedPath+internalDir[i]):
                        os.makedirs(DatedPath+internalDir[i])
            if not working_99_Exists:
                os.makedirs(DatedPath+'/99-Working')
                # os.makedirs(DatedPath+"/70-Packing-Slip")


        else:
            pass
            # logger.info("Folder '"+ClientCode+"-"+year +"/"+str(OrderDate)+"' exists.")

    except Exception as e:
        logger.error("Error while checking folder structure:  "+str(e))
        exc_type, exc_obj, exc_tb = sys.exc_info()
        fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
        print(exc_type, fname, exc_tb.tb_lineno)
        print("Error while checking folder structure:  "+str(e))


def mergeExcelsToOne(RootFolder, POSource, OrderDate, ClientCode, base_path):
    # converting str to datetime
    OrderDate = datetime.strptime(OrderDate, '%Y-%m-%d')
    # extracting year from the order date
    year = OrderDate.strftime("%Y")
    # formatting order date {2022-00-00) format
    OrderDate = OrderDate.strftime('%Y-%m-%d')
    inputpath = base_path + "/99-Working/40-Extract-Excel/"
    outputpath = base_path + "/50-Consolidate-Orders/"

    try:
        # logger.info('Checking 40-Extract-Excel directory exists or not.')
        file_list = glob.glob(inputpath + "/*.xlsx")
        if len(file_list) == 0:
            logger.info('No excel files found to merge.')
            print('No excel files found to merge.')
            return
        else:
            excl_list = []
            print("Merging files...")
            logger.info("Merging files...")
            # file_logger.info("Merging files...")
            for f in os.listdir(inputpath):
                # logger.info("Accessing '"+f+"' right now: ")
                # print("Accessing '"+f+"' right now: ")
                df = pd.read_excel(inputpath+"/"+f)
                df.insert(0, "file_name", f)
                excl_list.append(df)
            
            excl_merged = pd.concat(excl_list, ignore_index=True)
            #Loading Item Master
            with open(ConfigFolderPath, 'r') as jsonFile:
                config = json.load(jsonFile)
                df_item_master = pd.read_excel(ITEMMASTERPATH)
            
            # Renaming EAN to Article EAN for merge
            df_item_master.rename(columns={'EAN': 'ArticleEAN', 'MRP': 'Rate(Item Mster)'}, inplace=True)
            df_item_master = df_item_master[['ArticleEAN', 'Rate(Item Mster)', 'MRP Change Flag']]
        
            df_join_consolidate = excl_merged.merge(df_item_master, on= 'ArticleEAN', how="left")

            df_join_consolidate.to_excel(
                outputpath+"/"+'Consolidate-Orders.xlsx', index=False)
            logger.info("Merged "+str(len(file_list)) +
                        " excel files into a single excel file 'Consolidate-Orders.xlsx'")
            print("Merged "+str(len(file_list)) +
                  " excel files into a single excel file 'Consolidate-Orders.xlsx'")
            return 'All excels are merged into a single excel file'
    except Exception as e:
        logger.info("Error while merging files: "+str(e))
        exc_type, exc_obj, exc_tb = sys.exc_info()
        fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
        print(exc_type, fname, exc_tb.tb_lineno)
        print("Error while merging files: "+str(e))


def autoAllocation(workbook_path, workbook_sheet):
    try:
        start_cols = 3
        start_rows = 11
        logger.info('Auto Allocation of quantity based on Closing Stock is in progress...')
        req_sum_workbook = load_workbook(workbook_path)
        req_sum_sheet = req_sum_workbook[workbook_sheet]

        max_rows = req_sum_sheet.max_row
        max_cols = req_sum_sheet.max_column
        

        for i in range(start_rows, max_rows+1):
            cls_stk = 0
            grand_total = 0
            diff_gt_cs = 0
            cls_stk =  req_sum_sheet.cell(i,max_cols-2).value

            for j in range(start_cols, max_cols-3): # to calculate Grand Total of qty
                # print(req_sum_sheet.cell(i,j).value)
                if str(req_sum_sheet.cell(i,j).value).isnumeric(): # Removing Nonetype
                    grand_total = grand_total + req_sum_sheet.cell(i,j).value
            
            diff_gt_cs = cls_stk - grand_total # calculating diff between closing stock and grand total 
            # print(cls_stk, grand_total, diff_gt_cs)
            if diff_gt_cs < 0:
                for k in range(start_cols, max_cols-3):
                    if str(req_sum_sheet.cell(i,k).value).isnumeric():
                        if cls_stk >= int(str(req_sum_sheet.cell(i, k).value)):
                            cls_stk = cls_stk - int(str(req_sum_sheet.cell(i, k).value))
                        elif cls_stk == 0:
                            req_sum_sheet.cell(i, k).value = cls_stk
                        elif cls_stk < int(str(req_sum_sheet.cell(i, k).value)) and cls_stk > 0:
                            req_sum_sheet.cell(i, k).value = cls_stk
                            cls_stk = 0
            
        
        req_sum_workbook.save(workbook_path)
    except Exception as e:
        exc_type, exc_obj, exc_tb = sys.exc_info()
        fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
        print(exc_type, fname, exc_tb.tb_lineno)
        logger.error('Error while auto allocation of quantity'+str(e))



def mergeToPivotRQ(RootFolder, POSource, OrderDate, ClientCode, formulaWorksheet, TemplateFiles, base_path, reqSumTemplatePath):


    try:
        df_formula_cols = pd.read_excel(
            formulaWorksheet, sheet_name="Validate Column Names")

        # if not os.path.exists(RootFolder+"/"+ClientCode+"-"+year+"/"+OrderDate+"/50-Consolidate-Orders/Consolidate-Orders.xlsx"):
        if not os.path.exists(base_path + "/50-Consolidate-Orders/Consolidate-Orders.xlsx"):
            logger.info(
                "Could not find the consolidated order folder to generate requirement summary file")
            print(
                "Could not find the consolidated order folder to generate requirement summary file")
            return
        else:

            df_consolidated_order = pd.read_excel(
                base_path + "/50-Consolidate-Orders/Consolidate-Orders.xlsx")


            df_item_master = pd.read_excel(ITEMMASTERPATH)
            df_location_master = pd.read_excel(LOCATIONMASTERPATH)
            df_closing_stock = pd.read_excel(CLOSINGSTOCKMASTERPATH, skiprows=3)

            # --------------------
            # Renaming EAN as Article number to perform join using ArticleEAN
            df_item_master.rename(columns={'EAN': 'ArticleEAN'}, inplace=True)

            # common_columns = df_validate_item_master.merge(df_item_master, on= )

            # Merge on EAN from Item master
            df_SKU = df_consolidated_order.merge(df_item_master, on='ArticleEAN', how='left')

            df_SKU.rename(columns={'MRP_x': 'MRP'}, inplace=True)
            df_SKU_nodups = df_SKU.drop_duplicates()  # dropping duplicates

            df_SKU_nodups.to_excel(base_path + "/50-Consolidate-Orders/df_join_SKU.xlsx", columns=[
                                   'POItem','ArticleEAN', 'SKU', 'Qty', 'MRP', 'Receiving Location', 'Style', 'Style Name', 'PO Number'], index=False)
            # Opening df_SKU excel as df_SKU dataframe
            df_SKU = pd.read_excel(base_path + "/50-Consolidate-Orders/df_join_SKU.xlsx")



            # Perfoming join to get values of GST/ Allocation order
            df_gst_type = df_SKU.merge(df_location_master, on='Receiving Location', how='left')
            
            df_gst_type.to_excel(base_path + "/50-Consolidate-Orders/df_gst_type_merge_Receiving Location_left.xlsx")

            df_gst_type_nodups = df_gst_type.drop_duplicates()  # dropping duplicates

            df_gst_type.to_excel(base_path + "/50-Consolidate-Orders/df_join_gst_type.xlsx", index=False)

            # Opening df_gst_type excel as df_gst_type dataframe
            df_gst_type = pd.read_excel(base_path + "/50-Consolidate-Orders/df_join_gst_type.xlsx")

            # Perfoming join to get values of closing stock
            df_join_cl_stk = df_gst_type.merge(df_closing_stock, on='SKU', how='left')

            df_join_cl_stk_nodups = df_join_cl_stk.drop_duplicates()  # dropping duplicates

            df_join_cl_stk_nodups['Actual qty'] = df_join_cl_stk_nodups['Actual qty'].fillna(0)

            df_join_cl_stk_nodups.rename(columns={'Style_x': 'Style', 'Style Name_x': 'Style Name'}, inplace=True)


            df_join_cl_stk_nodups.to_excel(base_path + "/50-Consolidate-Orders/df_join_cl_stk.xlsx", index=False, 
            columns=['POItem','ArticleEAN', 'SKU', 'Qty', 'MRP','Receiving Location', 'Style', 'Style Name',	'PO Number', 'SGST/IGST Type','Allocation Order', 'Actual qty'])

            # Opening df_gst_type excel as df_gst_type dataframe
            df_join_cl_stk = pd.read_excel(base_path + "/50-Consolidate-Orders/df_join_cl_stk.xlsx")

            df_join_cl_stk['Order No.'] = ''  # adding order number as col
            df_join_cl_stk['Grand Total'] = ''  # adding Grand Total as col
            # print(df_join_cl_stk.tail(10))
            df_join_cl_stk['SGST/IGST Type'] = df_join_cl_stk['SGST/IGST Type'].fillna('---')
            df_join_cl_stk['Allocation Order'] = df_join_cl_stk['Allocation Order'].fillna('---')
            # print(df_join_cl_stk.tail(10))
            # final file used by requirement summary to make pivot
            df_join_cl_stk.to_excel(base_path + "/50-Consolidate-Orders/df_join_pivot.xlsx", index=False)
            
            df_pivot_final_join = pd.read_excel(base_path + "/50-Consolidate-Orders/df_join_pivot.xlsx")
            # --------------------

            # Constant Variables used in loops
            workbook_path = base_path + "/60-Requirement-Summary/Requirement-Summary.xlsx"
            workbook_sheet = 'Requirement Summary'
            color = "00FFCC99"
            red_color = "00FF0000"
            start_cols = 3
            start_rows = 11

            df_pivot = pd.pivot_table(df_join_cl_stk, index=["ArticleEAN", 'Actual qty', "SKU"], values='Qty',
                                      columns=['Allocation Order', 'PO Number', 'Order No.', 'Grand Total', 'SGST/IGST Type', 'Receiving Location'], aggfunc='sum')

            df_pivot['Grand Total'] = 0
            df_pivot['Closing Stock'] = 0
            df_pivot['Diff CS - GT'] = 0
            df_pivot['Rate'] = 0

            df_pivot.to_excel(workbook_path, sheet_name=workbook_sheet)

            # open pivot sheet again
            df_temp_p = pd.read_excel(workbook_path, sheet_name=workbook_sheet)
            df_temp = pd.read_excel(workbook_path, sheet_name=workbook_sheet, skiprows=6)
            df_temp.to_excel(base_path + "/50-Consolidate-Orders/"+"df_temp.xlsx",columns=['Actual qty'], index=False)
            tempWorkbook = load_workbook(base_path + "/50-Consolidate-Orders/"+"df_temp.xlsx")
            tempSheet = tempWorkbook.active
            # Inserting 5 rows to handle the gap between closing stock header and starting(entry) rows
            tempSheet.insert_rows(2, 6)
            # Saving closing stock and MRP in this temp sheet for later use
            tempWorkbook.save(base_path + "/50-Consolidate-Orders/"+"df_temp.xlsx")

            df_temp = pd.read_excel(base_path + "/50-Consolidate-Orders/"+"df_temp.xlsx")
            # Copying Closing stock from temp file to requirement summary file
            df_temp_p['Closing Stock'] = df_temp['Actual qty']
            print('Fetching Rate values from Item Master...')
            # Copying MRP from temp file to requirement summary file
            # df_temp_p['Rate'] = 0

            rq_template_source = REQSUMTEMPLATEPATH
            rq_summary_destination = base_path +"/60-Requirement-Summary/Temp-Requirement-Summary.xlsx"

            shutil.copy(rq_template_source, rq_summary_destination)

            with pd.ExcelWriter(rq_summary_destination, mode='a', engine='openpyxl', if_sheet_exists='replace') as rq_sum_writer:
                # Saving RQ Sum after adding CS, MRP
                df_temp_p.to_excel(
                    rq_sum_writer, sheet_name=workbook_sheet, index=False)

            pivotWorksheet = load_workbook(rq_summary_destination)
            pivotSheet = pivotWorksheet[workbook_sheet]
            # Deleting the cols for Actual Oty and MRP_y
            pivotSheet.delete_cols(2, 1)
            # inerting rows for date and cient name
            pivotSheet.insert_rows(1, 3)

            pivotSheet.cell(1, 1).value = 'ClientName'
            pivotSheet.cell(1, 1).font = Font(bold=True)
            pivotSheet.cell(1, 2).value = get_client_name(ClientCode)

            pivotSheet.cell(2, 1).value = 'Order Date'
            pivotSheet.cell(2, 1).font = Font(bold=True)
            pivotSheet.cell(2, 2).value = OrderDate
            pivotSheet.cell(4, 1).value = ''  # Removing Unnamed: 0

            

            rows = pivotSheet.max_row  # get max rows
            cols = pivotSheet.max_column  # get max rows

            for j in range(start_cols, cols-3):  # Rows Grand total
                pivotSheet.cell(start_cols+4, j).value = "=SUM("+openpyxl.utils.cell.get_column_letter(j)+str(start_rows)+":"+openpyxl.utils.cell.get_column_letter(j)+str(rows)+")"  # Grand Total on second last col
                pivotSheet.cell(start_cols+3, j).fill = PatternFill(start_color=color, end_color=color, fill_type="solid")  # Color to order field

            for i in range(start_rows, rows+1):  # Cols Grand total
                pivotSheet.cell(i, cols-3).value = "=SUM(B"+str(i)+":" + openpyxl.utils.cell.get_column_letter(cols-4)+str(i)+")"
                pivotSheet.cell(i, cols-1).value = "="+openpyxl.utils.cell.get_column_letter(cols-2)+str(i)+"-"+openpyxl.utils.cell.get_column_letter(cols-3)+str(i)

            # Applying Bold text
            pivotSheet.cell(5, 2).font = Font(bold=True) # PO Number
            pivotSheet.cell(6, 2).font = Font(bold=True) # Order No.
            pivotSheet.cell(7, 2).font = Font(bold=True) # Grand Total
            pivotSheet.cell(8, 2).font = Font(bold=True) # SGST/IGST Type
            pivotSheet.cell(9, 2).font = Font(bold=True) # Receiving Location
            pivotSheet.cell(10, 1).font = Font(bold=True) # ArticleEAN
            pivotSheet.cell(10, 2).font = Font(bold=True) # SKU

            thin_border = Border(left=Side(style='thin'),
                                 right=Side(style='thin'),
                                 top=Side(style='thin'),
                                 bottom=Side(style='thin'))

            for i in range(1, rows+1):
                for j in range(1, cols+1):
                    # if pivotSheet.cell(row=i, column=j).value == '---':
                    #     pivotSheet.cell(row=i, column=j).fill = PatternFill(start_color=red_color, end_color=red_color, fill_type="solid")
                    pivotSheet.cell(row=i, column=j).border = thin_border
                    pivotSheet.cell(row=i, column=j).alignment = Alignment(
                        horizontal='center', vertical='center')

            dim_holder = DimensionHolder(worksheet=pivotSheet)
            for col in range(pivotSheet.min_column, pivotSheet.max_column + 1):
                dim_holder[get_column_letter(col)] = ColumnDimension(
                    pivotSheet, min=col, max=col, width=20)
            pivotSheet.column_dimensions = dim_holder

            for r in range(11, rows+1):
                pivotSheet[f'A{r}'].number_format = '0'
            
            # For loop to remove unnamed values from allocation order row
            # counter = 0
            for j in range(3, cols-3):
                check_unnamed = str(pivotSheet.cell(4,j).value)
                if check_unnamed.__contains__('Unnamed'):
                    pivotSheet.cell(4,j).value = pivotSheet.cell(4,j-1).value
                    # counter +=1
                    # print(counter)

            # Making copy of Requirment Summary sheet
            Sheet2 = pivotWorksheet.copy_worksheet(pivotSheet)
            # Changing sheet name from 'Requirement Summary Copy' to 'Original Requirement Summary'
            Sheet2 = pivotWorksheet['Requirement Summary Copy']
            Sheet2.title  = 'Requirement Summary(Original)'
            # Saving workbook
            pivotWorksheet.save(workbook_path)
            # Removing temporary files
            os.remove(rq_summary_destination)
            os.remove(base_path + "/50-Consolidate-Orders/"+"df_join_cl_stk.xlsx")
            os.remove(base_path + "/50-Consolidate-Orders/"+"df_join_gst_type.xlsx")
            os.remove(base_path + "/50-Consolidate-Orders/"+"df_join_pivot.xlsx")
            os.remove(base_path + "/50-Consolidate-Orders/"+"df_join_SKU.xlsx")
            os.remove(base_path + "/50-Consolidate-Orders/"+"df_temp.xlsx")
            os.remove(base_path + "/50-Consolidate-Orders/df_gst_type_merge_Receiving Location_left.xlsx")

            # Auto Allocate Functionality###########################################################################################################
            with open(ConfigFolderPath, 'r') as jsonFile:
                config = json.load(jsonFile) 
                if config['autoAllocation'] == 'Y':
                    # print(config['autoAllocation'])
                    autoAllocation(workbook_path, workbook_sheet)


            print('Requirements summary sheet generated.')
            logger.info('Requirements summary sheet generated.')

            return 'Generated Requirement Summary file'

    except Exception as e:
        logger.error("Error while generating Requirement-Summary file: "+str(e))
        exc_type, exc_obj, exc_tb = sys.exc_info()
        fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
        print(exc_type, fname, exc_tb.tb_lineno)
        print("Error while generating Requirement-Summary file: "+str(e))









def validateRequirementSummary(InputSheet, cls_stk_column):
    try:
        logger.info("Validating Requirement Summary file")
        error = False
        # print(InputSheet.cell("B10").value)
        # if str(InputSheet.cell(4,2).value) == "Allocation Order" :#and InputSheet.cell(5,2) == "PO Number" and InputSheet.cell(6,2) == "Order No." and InputSheet.cell(7,2) == "Grand Total" and InputSheet.cell(8,2) == "SGST/IGST Type" and InputSheet.cell(9,2) == "Receiving Location" and InputSheet.cell(10,1) == "ArticleEAN" and InputSheet.cell(10,2) == "SKU" and InputSheet.cell(4,cls_stk_column-1) == "Grand Total" and InputSheet.cell(4,cls_stk_column) == "Closing Stock" and InputSheet.cell(4,cls_stk_column+1) == "Diff CS - GT" and InputSheet.cell(4,cls_stk_column+2) == "Rate":
        #     allocation_column = 2
        if str(InputSheet.cell(4,2).value) != "Allocation Order" :
            error = True
            logger.error("Please check Requirement Summary sheet, Allocation Order row not found")
        if str(InputSheet.cell(5,2).value) != "PO Number" :
            error = True
            logger.error("Please check Requirement Summary sheet, 'PO Number' row not found")
        if str(InputSheet.cell(6,2).value) != "Order No." :
            error = True
            logger.error("Please check Requirement Summary sheet, 'Order No.' row not found")
        if str(InputSheet.cell(7,2).value) != "Grand Total" :
            error = True
            logger.error("Please check Requirement Summary sheet, 'Grand Total' row not found")
        if str(InputSheet.cell(8,2).value) != "SGST/IGST Type" :
            error = True
            logger.error("Please check Requirement Summary sheet, 'SGST/IGST Type' row not found")
        if str(InputSheet.cell(9,2).value) != "Receiving Location" :
            error = True
            logger.error("Please check Requirement Summary sheet, 'Receiving Location' row not found")
        if str(InputSheet.cell(10,1).value) != "ArticleEAN" :
            error = True
            logger.error("Please check Requirement Summary sheet, 'ArticleEAN' column not found")
        if str(InputSheet.cell(10,2).value) != "SKU" :
            error = True
            logger.error("Please check Requirement Summary sheet, 'SKU' column not found")
        if str(InputSheet.cell(4,cls_stk_column-1).value) != "Grand Total" :
            error = True
            logger.error("Please check Requirement Summary sheet, 'Grand Total' column not found")
        if str(InputSheet.cell(4,cls_stk_column+1).value) != "Diff CS - GT" :
            error = True
            logger.error("Please check Requirement Summary sheet, 'Diff CS - GT' column not found")
        if str(InputSheet.cell(4,cls_stk_column+2).value) != "Rate" :
            error = True
            logger.error("Please check Requirement Summary sheet, 'Rate' column not found")
        
        return error
    except Exception as e:
        logger.error("Error while validating requirement summary: "+ str(e))
        exc_type, exc_obj, exc_tb = sys.exc_info()
        fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
        print(exc_type, fname, exc_tb.tb_lineno)
        return error
    




def generatingPackingSlip(RootFolder, ReqSource, OrderDate, ClientCode, formulaWorksheet, TemplateFiles, base_path, packingSlipTemplatePath):
    try:
        # converting str to datetime
        OrderDate = datetime.strptime(OrderDate, '%Y-%m-%d')
        # extracting year from the order date
        year = OrderDate.strftime("%Y")
        # formatting order date {2022-00-00) format
        OrderDate = OrderDate.strftime('%Y-%m-%d')

        startedTemplating = time.time()
        sourcePivot = ReqSource
        source = PACKINGSLIPTEMPLATEPATH
        destination = base_path + "/99-Working/TemplateFile.xlsx"

        # Making Copy of template file
        # shutil.copy(source, destination)
        # print("File copied successfully.")

        # Load work vook and sheets
        InputWorkbook = load_workbook(sourcePivot, data_only=True)
        # TemplateWorkbook = load_workbook(destination)

        InputSheet = InputWorkbook['Requirement Summary']
        # TemplateSheet = TemplateWorkbook['ORDER']
        cls_stk_column = 0
        for j in range(1, 10000):
                if str(InputSheet.cell(4, j).value) == 'Closing Stock':
                    cls_stk_column = j
                    break
        
        allocation_column = 2

        error = validateRequirementSummary(InputSheet, cls_stk_column)
        if error == True: 
            return
            

        # Get rows and Column count
        df = pd.DataFrame(InputSheet, index=None)
        rows = len(df.axes[0])
        # cols = len(df.axes[1])
        cols = cls_stk_column + 2
        
        # print(formulaWorksheet)
        formulaWorksheet = load_workbook(formulaWorksheet, data_only=True)
        formulaSheet = formulaWorksheet['FormulaSheet']
        DBFformula = formulaWorksheet['DBF']



        # df_SGST = pd.read_excel(SGSTMASTERPATH,sheet_name='DBF')
        print('Loading Master files for processing...')
        logger.info('Loading Master files for processing...')

        # Opening Item Master Sheet
        df_item_master = pd.read_excel(ITEMMASTERPATH, sheet_name='Item Master', index_col=False)
        print('Item Master loaded.')
        logger.info('Item Master loaded.')
        # Opening IGST Master Sheet
        df_IGSTMaster = pd.read_excel(IGSTMASTERPATH, sheet_name='DBF', index_col=False)
        print('IGST Master loaded.')
        logger.info('IGST Master loaded.')
        # Opening SGST Master Sheet
        df_SGSTMaster = pd.read_excel(SGSTMASTERPATH, sheet_name='DBF', index_col=False)
        print('SGST Master loaded.')
        logger.info('SGST Master loaded.')
        # Opening Location2 Master Sheet
        df_Location2 = pd.read_excel(LOCATION2MASTERPATH, sheet_name='Location2', index_col=False)
        print('Location 2 Master loaded.')
        logger.info('Location 2 Master loaded.')


        start_cols = allocation_column + 1


        for column in range(start_cols, cols-3):
            startedTemplatingFile = time.time()
            # Making Copy of template file
            shutil.copy(source, destination)
            # logger.info("Template File copied successfully for generating packing-slip")
            # print("Template File copied successfully for generating packing-slip")

            # Load work vook and sheets
            TemplateWorkbook = load_workbook(destination)
            TemplateSheet = TemplateWorkbook['ORDER']
            PackingSheet = TemplateWorkbook['Packing Slip']
            dbfsheet = TemplateWorkbook['DBF']
            
        

            TemplateSheet.cell(5, 1).value = InputSheet.cell(start_cols+3, column).value  # Order Name

            # PO Number
            filename = str(TemplateSheet.cell(5, 1).value)
            if filename == '' or filename == None:
                logger.error("Order number field is empty. Please check Requirement Summary.")
                return
            # print(filename)
            filename = "".join(x for x in filename if x.isalnum()) # this handles and reomoves special character(!@#$%^&*()_+{}:"|<>?")
            TemplateSheet.cell(5, 2).value = InputSheet.cell(start_cols+2, column).value

            #

            # Receving Location
            TemplateSheet.cell(5, 3).value = InputSheet.cell(start_cols+6, column).value

            TemplateSheet.cell(1, 1).value = 'Order Date'
            TemplateSheet.cell(1, 2).value = InputSheet.cell(2, 2).value  # Date

            TemplateSheet.cell(1, 3).value = 'SGST/IGST'
            TemplateSheet.cell(1, 4).value = InputSheet.cell(start_cols+5, column).value  # IGST/SGST Type (6,cols-3)
            if TemplateSheet.cell(1, 4).value == None or TemplateSheet.cell(1, 4).value == '---':
                print(
                    "IGST/SGST TYPE is not found for order number "+filename+" ! Please check the Requirment Summary file and process again.")
                logger.error(
                    "IGST/SGST TYPE is not found for order number "+filename+" ! Please check the Requirment Summary file and process again.")
                break
                return
                

            # Copy EAN to template sheet
            Trows = 8
            InRows = 11
            Tcols = 5
            dbfrows = 2
            dbfcols = 57
            dbf_start_col = 1
            for row in range(InRows, rows+1):
                # if InputSheet.cell(row,column).value != None or InputSheet.cell(row,column).value != "":
                if str(InputSheet.cell(row, column).value).isnumeric() and str(InputSheet.cell(row, column).value) != "0":

                    # Copy Qty to template sheet
                    TemplateSheet.cell(Trows, Tcols).value = InputSheet.cell(row, column).value
                    TemplateSheet.cell(Trows, Tcols+1).value = InputSheet.cell(row, column).value
                    TemplateSheet.cell(Trows, Tcols+2).value = "="+openpyxl.utils.cell.get_column_letter(Tcols+1)+str(Trows)  # Actual Qty in packing(order sheet)
                    TemplateSheet.cell(Trows, Tcols+3).value = InputSheet.cell(row, cols).value # Rate from Requirement summry

                    # Copy EAN to template sheet
                    TemplateSheet.cell(Trows, Tcols-3).value = InputSheet.cell(row, 1).value

                    # VLOOKUP - formula sheet
                    # StyleName
                    
                    TemplateSheet.cell(Trows,Tcols-4).value = "="+formulaSheet.cell(3,2).value.replace("#VAL#",str(Trows-6))
                    #TemplateSheet.cell(Trows, Tcols-4).value = "='Hidden Item Master'!A"+str(Trows-6)

                    # style
                    TemplateSheet.cell(Trows,Tcols-2).value =  "="+formulaSheet.cell(4,2).value.replace("#VAL#",str(Trows-6))
                    #TemplateSheet.cell(Trows, Tcols-2).value = "='Hidden Item Master'!C"+str(Trows-6)

                    # SADM SKU
                    TemplateSheet.cell(Trows,Tcols-1).value = "="+formulaSheet.cell(5,2).value.replace("#VAL#",str(Trows-6))
                    #TemplateSheet.cell(Trows, Tcols-1).value = "='Hidden Item Master'!D"+str(Trows-6)

                    # MRP Change Flag
                    TemplateSheet.cell(Trows, Tcols+19).value = "="+formulaSheet.cell(6,2).value.replace("#VAL#",str(Trows-6))
                    # TemplateSheet.cell(Trows, Tcols+19).value = "=IF('Hidden Item Master'!H"+str(Trows-6)+'<>"","*","")'
                    


                    # Rate (in Rs.) Order file
                    # TemplateSheet.cell(Trows,Tcols+3).value = "="+formulaSheet.cell(6,2).value.replace("#VAL#",str(Trows))
                    # TemplateSheet.cell(Trows, Tcols+3).value = "='Hidden Item Master'!E"+str(Trows-6)

                    # Cls stk vs order
                    #TemplateSheet.cell(Trows,Tcols+6).value = TemplateSheet.cell(Trows,Tcols+5).value - TemplateSheet.cell(Trows,Tcols+2).value
                    TemplateSheet.cell(Trows, Tcols+6).value = "="+openpyxl.utils.cell.get_column_letter(Tcols+5)+str(Trows) + '-'+openpyxl.utils.cell.get_column_letter(Tcols+2)+str(Trows)

                    # LOCATION2
                    TemplateSheet.cell(Trows,Tcols+7).value = "="+formulaSheet.cell(8,2).value.replace("#VAL#",str(Trows-6))
                    #TemplateSheet.cell(Trows, Tcols+7).value = "='Hidden Item Master'!F"+str(Trows-6)

                    # BULK  / DTA  BULK  /  EOSS LOC
                    TemplateSheet.cell(Trows,Tcols+8).value = "="+formulaSheet.cell(9,2).value.replace("#VAL#",str(Trows-6))
                    #TemplateSheet.cell(Trows, Tcols+8).value = "='Hidden Item Master'!G"+str(Trows-6)

                    # MRP
                    TemplateSheet.cell(Trows,Tcols+9).value = "="+formulaSheet.cell(10,2).value.replace("#VAL#",str(Trows-6))
                    #TemplateSheet.cell(Trows, Tcols+9).value = "='Hidden Item Master'!E"+str(Trows-6)

                    # Closing stk
                    # TemplateSheet.cell(Trows,Tcols+5).value = "="+formulaSheet.cell(11,2).value.replace("#VAL#",str(Trows))
                    TemplateSheet.cell(Trows, Tcols+5).value = InputSheet.cell(row, cols-2).value

                    # SCAN
                    TemplateSheet.cell(Trows,Tcols+10).value = "="+formulaSheet.cell(11,2).value.replace("#VAL#",str(Trows))
                    # TemplateSheet.cell(Trows, Tcols+10).value = "=SUMIF('packing slip'!$B$3:$B$999,$B"+str(Trows)+",'packing slip'!$C$3:$C$999)"

                    # SCAN VS DIFF
                    TemplateSheet.cell(Trows,Tcols+11).value = "="+formulaSheet.cell(12,2).value.replace("#VAL#",str(Trows))
                    # TemplateSheet.cell(Trows, Tcols+11).value = "=O"+str(Trows)+"-G"+str(Trows)

                    # ERROR
                    TemplateSheet.cell(Trows,Tcols+12).value = "="+formulaSheet.cell(13,2).value.replace("#VAL#",str(Trows))
                    # TemplateSheet.cell(Trows, Tcols+12).value = "=IF(P"+str(Trows)+"<>0,1,0)"

                    # STYLE COLOR
                    TemplateSheet.cell(Trows,Tcols+13).value = "="+formulaSheet.cell(14,2).value.replace("#VAL#",str(Trows))
                    # TemplateSheet.cell(Trows, Tcols+13).value = "=IF(P"+str(Trows)+"<=0,0,1)"

                    # Adding values to DBF
                    for j in range(dbf_start_col, dbfcols):
                        dbfsheet.cell(dbfrows, j).value = '='+DBFformula.cell(2, j+1).value.replace(
                            "#VAL#", str(Trows)).replace("#DBFROWS#", str(dbfrows))

                    Trows += 1
                    dbfrows += 1

            TemplateSheet.cell(5, 5).value = "=SUM(E8:E"+str(Trows-1)+")"
            TemplateSheet.cell(5, 6).value = "=SUM(F8:F"+str(Trows-1)+")"
            TemplateSheet.cell(5, 7).value = "=SUM(G8:G"+str(Trows-1)+")"

            TemplateWorkbook.save(base_path + "/70-Packing-Slip/"+""+str(filename)+".xlsx")
            TemplateWorkbook.close()

            # Opening Packing slip using openpyxl to check igst/sgst value
            sourcePackingSlip = base_path + "/70-Packing-Slip/" +str(filename)+".xlsx"
            TemplateWorkbook = load_workbook(sourcePackingSlip, data_only=True)
            TemplateSheet = TemplateWorkbook['ORDER']
            # print(TemplateSheet.cell(1,4).value)
            # Getting data from IGST/ SGST Sheet DBF to Tempalate sheet DBF

            # If IGST then open IGST Master
            if TemplateSheet.cell(1, 4).value == 'IGST':

                # Opening Packing slip as df for second time to get EAN values
                df_TempalateWorkbook = pd.read_excel(sourcePackingSlip, sheet_name='ORDER', skiprows=6, index_col=False)
                # df_TempalateWorkbook.rename(columns = {'EAN':'EAN ID'}, inplace = True)

                # Temporary df to store EAN
                df_EAN_temp = df_TempalateWorkbook[['EAN']]
                # Applying join using EAN ID to get hidden_item_master to geSKU IDt SKU and Other fields
                df_hidden_item_master = df_EAN_temp.merge(df_item_master, on='EAN', how='left')

                df_Location2_temp = df_hidden_item_master.merge(df_Location2, on='EAN', how='left')
                # print(df_Location2_temp.head(10))
                df_Location2_temp.rename(columns={'SKU_x': 'SKU'}, inplace=True)

                # Temporary df to store SKU
                df_SKU_temp = df_hidden_item_master[['SKU']]
                df_SKU_temp.rename(columns={'SKU': 'ITEMNAME'}, inplace=True)
                # Applying join using ITEMNAME to get hidden_dbf (from IGST/SGST sheet) to get SKU and Other fields
                df_GST_hidden = df_SKU_temp.merge(df_IGSTMaster, on='ITEMNAME', how='left')


                with pd.ExcelWriter(sourcePackingSlip, mode='a', engine='openpyxl', if_sheet_exists='replace') as writer:
                    # Adding Hidden Item Master sheet to the RQ sheet with values 'Style Name', 'EAN', 'Style', 'SKU', 'MRP'
                    df_Location2_temp.to_excel(writer, sheet_name='Hidden Item Master', index=False, columns=[
                    'Style Name', 'EAN', 'Style', 'SKU', 'MRP','Location 2', 'BULK  / DTA  BULK  /  EOSS LOC','MRP Change Flag'])
                    # # Adding Hidden DBF sheet to the RQ sheet with values 'Vouchertypename', 'CSNNO','DATE' ETC.
                    df_GST_hidden.to_excel(writer, sheet_name='Hidden DBF', index=False, columns=['Vouchertypename', 'CSNNO', 'DATE',
                                                                                                  'REFERENCE', 'REF1', 'DEALNAME', 'PRICELEVEL', 'ITEMNAME', 'GODOWN', 'QTY', 'RATE', 'SUBTOTAL', 'DISCPERC',
                                                                                                  'DISCAMT', 'ITEMVALUE', 'LedgerAcct', 'CATEGORY1', 'COSTCENT1', 'CATEGORY2', 'COSTCENT2', 'CATEGORY3', 'COSTCENT3',
                                                                                                  'CATEGORY4', 'COSTCENT4', 'ITEMTOTAL', 'TOTALQTY', 'CDISCHEAD', 'CDISCPERC', 'COMMONDISC', 'BEFORETAX',
                                                                                                  'TAXHEAD', 'TAXPERC', 'TAXAMT', 'STAXHEAD', 'STAXPERC', 'STAXAMT', 'ITAXHEAD', 'ITAXPERC', 'ITAXAMT', 'NETAMT',
                                                                                                  'ROUND', 'ROUND1', 'REFTYPE', 'Name', 'REFAMT', 'Narration', 'Transport', 'transmode', 'pymtterm', 'ordno',
                                                                                                  'orddate', 'DANO', 'Delyadd1', 'Delyadd2', 'Delyadd3', 'Delyadd4'])

            # If SGST then open SGST Master
            if TemplateSheet.cell(1, 4).value == 'SGST':

                # Opening Packing slip as df for second time to get EAN values
                df_TempalateWorkbook = pd.read_excel(sourcePackingSlip, sheet_name='ORDER', skiprows=6, index_col=False)
                # df_TempalateWorkbook.rename(columns = {'EAN':'EAN ID'}, inplace = True)

                # Temporary df to store EAN ID
                df_EAN_temp = df_TempalateWorkbook[['EAN']]
                # Applying join using EAN ID to get hidden_item_master to get SKU and Other fields

                df_hidden_item_master = df_EAN_temp.merge(df_item_master, on='EAN', how='left')

                df_Location2_temp = df_hidden_item_master.merge(df_Location2, on='EAN', how='left')
                # print(df_Location2_temp)
                df_Location2_temp.rename(columns={'SKU_x': 'SKU'}, inplace=True)

                # Temporary df to store SKU
                df_SKU_temp = df_hidden_item_master[['SKU']]
                df_SKU_temp.rename(columns={'SKU': 'ITEMNAME'}, inplace=True)
                # Applying join using ITEMNAME to get hidden_dbf (from IGST/SGST sheet) to get SKU and Other fields
                df_GST_hidden = df_SKU_temp.merge(df_SGSTMaster, on='ITEMNAME', how='left')
                # for col in df_Location2_temp.columns:
                #     print(col)
                with pd.ExcelWriter(sourcePackingSlip, mode='a', engine='openpyxl', if_sheet_exists='replace') as writer:
                    # Adding Hidden Item Master sheet to the RQ sheet with values 'Style Name', 'EAN', 'Style', 'SKU', 'MRP'
                    df_Location2_temp.to_excel(writer, sheet_name='Hidden Item Master', index=False, columns=
                    ['Style Name', 'EAN', 'Style', 'SKU', 'MRP','Location 2', 'BULK  / DTA  BULK  /  EOSS LOC','MRP Change Flag'])
                    # # Adding Hidden DBF sheet to the RQ sheet with values 'Vouchertypename', 'CSNNO','DATE' ETC.
                    df_GST_hidden.to_excel(writer, sheet_name='Hidden DBF', index=False, columns=[
                        'Vouchertypename', 'CSNNO', 'DATE','REFERENCE', 'REF1', 'DEALNAME', 'PRICELEVEL', 'ITEMNAME', 
                        'GODOWN', 'QTY', 'RATE', 'SUBTOTAL', 'DISCPERC','DISCAMT', 'ITEMVALUE', 'LedgerAcct', 'CATEGORY1', 
                        'COSTCENT1', 'CATEGORY2', 'COSTCENT2', 'CATEGORY3', 'COSTCENT3','CATEGORY4', 'COSTCENT4', 'ITEMTOTAL', 
                        'TOTALQTY', 'CDISCHEAD', 'CDISCPERC', 'COMMONDISC', 'BEFORETAX','TAXHEAD', 'TAXPERC', 'TAXAMT', 'STAXHEAD', 
                        'STAXPERC', 'STAXAMT', 'ITAXHEAD', 'ITAXPERC', 'ITAXAMT', 'NETAMT','ROUND', 'ROUND1', 'REFTYPE', 'Name', 
                        'REFAMT', 'Narration', 'Transport', 'transmode', 'pymtterm', 'ordno','orddate', 'DANO', 'Delyadd1', 'Delyadd2', 
                        'Delyadd3', 'Delyadd4'])
                pass

            # Full Item Master
            with pd.ExcelWriter(sourcePackingSlip, mode='a', engine='openpyxl', if_sheet_exists='replace') as writer:
                df_item_master.to_excel(writer, sheet_name='Hidden Full Item Master', index=False, columns=['Style Name', 'EAN', 'Style', 'SKU', 'MRP','MRP Change Flag'])

            TemplateWorkbook.close()

            # Opening Sheet again to hide hidden dbf and item master sheets
            TemplateWorkbook = load_workbook(sourcePackingSlip)
            hidden_item_master = TemplateWorkbook['Hidden Item Master']
            hidden_dbf = TemplateWorkbook['Hidden DBF']
            hidden_full_item_master = TemplateWorkbook['Hidden Full Item Master']

            hidden_item_master.sheet_state = 'hidden'
            hidden_full_item_master.sheet_state = 'hidden'
            hidden_dbf.sheet_state = 'hidden'

            TemplateWorkbook.close()
            TemplateWorkbook.save(sourcePackingSlip)

            logger.info("Packing slip generated for order number "+str(filename) +
                        " in {:.2f}".format(time.time() - startedTemplatingFile, 2) + " seconds.")
            # file_logger.info("Packing slip generated for: "+str(filename)+ " file in {:.2f}".format(time.time() - startedTemplatingFile,2)+ " seconds.")
            print("Packing slip generated for: "+str(filename) +
                  " file in {:.2f}".format(time.time() - startedTemplatingFile, 2) + " seconds.")

        logger.info("Total time taken for generation of packing-slips:  {:.2f}".format(
            time.time() - startedTemplating, 2) + " seconds.")
        print("Total time taken for generation of packing-slips:  {:.2f}".format(
            time.time() - startedTemplating, 2) + " seconds.")
        return 'Completed!'
    except Exception as e:
        logger.error("Error while generating packing-slip file: "+str(e))
        exc_type, exc_obj, exc_tb = sys.exc_info()
        fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
        print(exc_type, fname, exc_tb.tb_lineno)
        print("Error while generating packing-slip file: "+str(e))
