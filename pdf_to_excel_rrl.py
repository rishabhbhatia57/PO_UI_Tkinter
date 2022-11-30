import re, os, time, sys
import parse
import pdfplumber
import pandas as pd
from collections import namedtuple
from PyPDF2 import PdfWriter, PdfReader
from PyPDF2 import PdfFileReader
import logging
import tabula
from datetime import datetime
import numpy as np
from openpyxl import load_workbook, Workbook
import xlsxwriter

import BKE_log

logger1 = logging.getLogger("PyPDF2")
logger1.setLevel(logging.ERROR)

logger = BKE_log.setup_custom_logger('root')


# For Reliance Reatial Limited

def pdfToTable_RRL(RootFolder, POSource, OrderDate, ClientCode, f, base_path):
    file_name=f
    po_number = ''
    Receving_location = ''
    try:
        # converting str to datetimedatetime
        OrderDate = datetime.strptime(OrderDate, '%Y-%m-%d')
        # extracting year from the order date
        year = OrderDate.strftime("%Y")
        # formatting order date {2022-00-00) format
        OrderDate = OrderDate.strftime('%Y-%m-%d')

        startedProcessing = time.time()
        # Getting data from 1 page
        # PO Number, receving Location

        with pdfplumber.open(base_path+'/99-Working/10-Download-Files/'+str(f)) as pdf:
            
            pages = pdf.pages[0]
            linesData = []
            # for page in pdf.pages:
            text = pages.extract_text()
            # print (text)
            for line in text.split('\n'):
                # print(line)
                #  or line.__contains__('SELLER') # line.__contains__('GOODS') or 
                if line.__contains__('REGISTERED OFFICE') or line.__contains__('computer generated') or \
                    line.__contains__('OWNER') or line.__contains__('CIN') or \
                    line.__contains__('Telephone:') or line.__contains__('Date') or line.__contains__('throughSRMPortal.Failureto')\
                    or line.__contains__('GSTINNo') or line.__contains__('Pin-Code') or line.__contains__('CourtHouse')\
                    or line.__contains__('VendorStatus') or line.__contains__('DeliveryAddress') or line.__contains__('Payment Terms')\
                    or line.__contains__('Thisisacomputergenerateddocumentnot') or line.__contains__('Tel:/Fax:') or line.__contains__('Phone'):  
                    pass
                else:
                    # print("#####################\n" , line , "\n######################")
                    # Logic for using PDFPlumber and Tabula based on "Purchase Order -Preview" or "Purchase Order"
                    # if line.__contains__('SELLER') and line.__contains__('PURCHASE') and line.__contains__('ORDER'):
                    if line.__contains__('SELLER PURCHASE ORDER') or line.__contains__('SELLER PURCHASEORDER'):
                        order_type = 'SELLER PURCHASE ORDER'
                    # elif line.__contains__('SELLER') and line.__contains__('PURCHASE') and line.__contains__('ORDER') and line.__contains__('PREVIEW'):
                    if line.__contains__('SELLER PURCHASE ORDER -PREVIEW') or line.__contains__('SELLER PURCHASEORDER -PREVIEW'):
                        order_type = 'SELLER PURCHASE ORDER -PREVIEW'
                    # else:
                    #     logger.error("PDF format not valid. Unable to extract data from PDF for file: " + file_name)
                    #     return


                    
                    if line.__contains__('PONo') or line.__contains__('PO NO') or line.__contains__('PO No') :
                        # print(line)
                        linesData.append(line.split(' '))
                        if order_type == 'SELLER PURCHASE ORDER -PREVIEW':
                            if line.__contains__('PONo'):
                                po_number = str(linesData[0][1]).replace('PONo.:', '')
                            if line.__contains__('PO No'):
                                temp_line_data = str(linesData[0][2])
                                index_site = temp_line_data.find('Site')
                                po_number = temp_line_data[0:index_site].replace('No.:', '')
                        if order_type == 'SELLER PURCHASE ORDER':
                            po_number = str(linesData[0][15])

                    # get Receving location
                    if line.__contains__('RRL'):
                        # print(line)
                        index1 = line.find('RRL')
                        index2 = line.find('SELLER')
                        Receving_location = str(line[index1:index2])
                        # print(Receving_location)



                    
        # print("\n")
        # print(po_number,Receving_location)

        # Excract EAN Data from Page 2 to len(Pages of PDF)
        if order_type == "SELLER PURCHASE ORDER -PREVIEW":
            # print("\nRunning PDF PLUMBER code\n") 
            with pdfplumber.open(base_path+'/99-Working/10-Download-Files/'+str(f)) as pdf:
                pages = pdf.pages
                # print(pages)
                pages_to_keep = []
                for i in range(1,len(pages)+1):
                    pages_to_keep.append(i)
                
                infile = PdfReader(base_path+'/99-Working/10-Download-Files/'+str(f), 'rb')
                output = PdfWriter()

                for i in range(1,len(pages_to_keep)):
                    # print(i)
                    p = infile.pages[i] 
                    output.add_page(p)

                with open(base_path+'/99-Working/10-Download-Files/'+str(f), 'wb') as f:
                    output.write(f)
                
                # file_path = base_path+'/99-Working/10-Download-Files/'+str(f)
                # print('\n',str(f),'\n',str(file_path),'\n')
                with pdfplumber.open(base_path+'/99-Working/10-Download-Files/'+str(file_name)) as pdf:
                    pages = pdf.pages
                    linesData = []
                    for page in pdf.pages:
                        text = page.extract_text()
                        for line in text.split('\n'):
                            if line.__contains__('DRAFT') or line.__contains__('             ') or line.__contains__('HSN/SACCode') or line.__contains__('VendorItemNo') or line.__contains__('PageNo') or line.__contains__('MaterialDescription') \
                                or line.__contains__('SubTotal') or line.__contains__('______________________________________________________________________________________________________________'):
                                pass
                            else:
                            # print(line.split(' '))
                                linesData.append(line.split(' '))
                combinedlinesData = []
                for i in range(0,len(linesData),3):
                    List1 = linesData[i]
                    List2 = linesData[i+1]
                    List3 = linesData[i+2]
                    combinedlinesData.append(List1)
                # print(List1,List2,List3)
                # print(combinedlinesData)
                # print(len(combinedlinesData))

                df = pd.DataFrame(combinedlinesData,columns=['POItem','Article Number','ArticleEAN','ArticleDescription','Qty', 'UM', 'MRP', 'BasicCostPrice(TaxableValue)','IGSTRate','IGSTAmt','Total Amount'])

                # removing last rows
                df.drop(df.tail(1).index, inplace=True)
                rows = len(df.axes[0])
                cols = len(df.axes[1])

                # Calculating grand total
                grand_total_text = str(df.iloc[rows-1]['POItem']).replace('_', '')
                # df.A.str.extract('(\d+)')
                grand_total_text = ''.join(filter(lambda i: i.isdigit(),grand_total_text))
                df.drop(df.tail(1).index, inplace=True)
                # Removing unwanted characters (,) in numeric values

                df['BasicCostPrice(TaxableValue)'] = df['BasicCostPrice(TaxableValue)'].str.replace(',', '')
                df['IGSTRate'] = df['IGSTRate'].str.replace(',', '')
                df['IGSTAmt'] = df['IGSTAmt'].str.replace(',', '')
                df['MRP'] = df['MRP'].str.replace(',', '')
                df['Total Amount'] = df['Total Amount'].str.replace(',', '')

                # Converting str to number
                df['POItem'] = pd.to_numeric(df['POItem'])
                df['Article Number'] = pd.to_numeric(df['Article Number'])
                df['Qty'] = pd.to_numeric(df['Qty'])
                df['ArticleEAN'] = pd.to_numeric(df['ArticleEAN'])
                df['MRP'] = pd.to_numeric(df['MRP'])
                df['BasicCostPrice(TaxableValue)'] = pd.to_numeric(df['BasicCostPrice(TaxableValue)'])
                df['IGSTRate'] = pd.to_numeric(df['IGSTRate'])
                df['IGSTAmt'] = pd.to_numeric(df['IGSTAmt'])
                df['Total Amount'] = pd.to_numeric(df['Total Amount'])
                df['Receiving Location'] = Receving_location
                df['PO Number'] = po_number
                df['PO Number'] = pd.to_numeric(df['PO Number'])

                df = df[['POItem','ArticleEAN','Article Number','ArticleDescription','MRP','BasicCostPrice(TaxableValue)','Qty', 'UM','Total Amount','IGSTRate','IGSTAmt','Receiving Location','PO Number']]
                # Inserting extra cols 
                df.insert(loc = 4,column = 'HSNCode',value = '')
                # df.insert(loc = 8,column = 'UM',value = '')
                df.insert(loc = 9,column = 'TaxableValue',value = '')
                df.insert(loc = 11,column = 'CGSTRate',value = '')
                df.insert(loc = 12,column = 'CGSTAmt',value = '')
                df.insert(loc = 13,column = 'SGSTRate',value = '')
                df.insert(loc = 14,column = 'SGSTAmt',value = '')
                df.insert(loc = 15,column = 'UTGSTRate',value = '')
                df.insert(loc = 16,column = 'UTGSTAmt',value = '')
                df.insert(loc = 19,column = 'Vendor Penalty',value = '')
                df.insert(loc = 20,column = 'Vendor Name',value = '')
                # df.insert(loc = 23,column = 'Vendor Code',value = '')




                Sum_of_Qty = df['Qty'].sum()
                # print(rows,cols)
                # print(grand_total_text,Sum_of_Qty)
                if int(grand_total_text) != int(Sum_of_Qty):
                    print('Grand Total of the converted excel does not match with the pdf.')
                    logger.info('Grand Total of the converted excel does not match with the pdf.')
                    return
                else:
                    df.to_excel(base_path+"/99-Working/40-Extract-Excel/"+po_number+'.xlsx',index=False)
                    print("Converted '" + file_name + "' to '"+po_number +".xlsx' in " + "{:.2f}".format(time.time() - startedProcessing, 2) + " seconds.")
                    logger.info("Converted '" + file_name + "' to '"+po_number +".xlsx' in " + "{:.2f}".format(time.time() - startedProcessing, 2) + " seconds.")

        if order_type == "SELLER PURCHASE ORDER":
            # print("\nRunning TABULA code\n")
            intermediateCSV = base_path+'/99-Working/30-Extract-CSV/'+str(f).replace('pdf', 'csv')
            intermediateExcel = base_path+'/99-Working/20-Intermediate-Files/'+str(f).replace('pdf', 'xlsx')


            tabula.convert_into(input_path=base_path+'/99-Working/10-Download-Files/'+str(f), output_path=intermediateCSV, pages= "all", lattice=True)
            
            df_rrl = pd.read_csv(filepath_or_buffer=intermediateCSV, skiprows= 8)

            writer = pd.ExcelWriter(intermediateExcel, engine='xlsxwriter')
            df_rrl.to_excel(writer, sheet_name='Sheet1', merge_cells=False, index=False)
            writer.save()

            workbook = load_workbook(filename=intermediateExcel)
            sheet = workbook.active

            max_rows = sheet.max_row

            for i in range(2, max_rows+1):
                if sheet.cell(i,1).value == 'Sub Total of Qty' or sheet.cell(i,1).value == 'Sr.No' or sheet.cell(i,1).value == '' or sheet.cell(i,1).value == None:
                    sheet.cell(i,1).value = "Remove"

            counter = max_rows+1
            while counter > 0:
                if sheet.cell(counter, 1).value == "Remove":
                    sheet.delete_rows(counter)
                counter = counter - 1
            
            max_rows = sheet.max_row
            max_cols = sheet.max_column
            for i in range(1, max_rows):
                for j in range(1, max_cols):
                    # print(sheet.cell(i,j).value)
                    if sheet.cell(i,j).value != None and sheet.cell(i,j).value != '' and '_x000D_' in sheet.cell(i,j).value:
                        temp = sheet.cell(i,j).value
                        temp_arr = temp.split("_x000D_")
                        sheet.cell(i,j).value = temp_arr[0]
            # print("max_rows: " + str(max_rows))
            grand_total_text = str(sheet.cell(max_rows-1,2).value)
            grand_total_text2 = grand_total_text.split('.')
            grand_total_text = grand_total_text2[0]
            # print('################## Grand Total: ' + grand_total_text)

            workbook.save(intermediateExcel)
            df = pd.read_excel(intermediateExcel, sheet_name= 'Sheet1')
            # print(df.head(10))

            df.rename(columns= {'Sr.No':'POItem', 'Article No.':'Article Number', 'EAN No.':'ArticleEAN', 'Material Description':'ArticleDescription','Quantity': 'Qty', 'UOM':'UM', 'Base Cost':'BasicCostPrice(TaxableValue)', 'IGST (%)':'IGSTRate', 'IGST':'IGSTAmt', 'Total Base Value':'Total Amount'}, inplace=True)
            # print(df.head(10))
            df.drop(df.tail(1).index, inplace=True)
            rows = len(df.axes[0])
            cols = len(df.axes[1])

            # Removing unwanted characters (,) in numeric values
            df['BasicCostPrice(TaxableValue)'] = df['BasicCostPrice(TaxableValue)'].str.replace(',', '')
            # df['IGSTRate'] = df['IGSTRate'].str.replace(',', '')
            # df['IGSTAmt'] = df['IGSTAmt'].str.replace(',', '')
            df['MRP'] = df['MRP'].str.replace(',', '')
            df['Total Amount'] = df['Total Amount'].str.replace(',', '')


            # Converting str to number
            df['POItem'] = pd.to_numeric(df['POItem'])
            df['Article Number'] = pd.to_numeric(df['Article Number'])
            df['Qty'] = pd.to_numeric(df['Qty'])
            df['ArticleEAN'] = pd.to_numeric(df['ArticleEAN'])
            df['MRP'] = pd.to_numeric(df['MRP'])
            df['BasicCostPrice(TaxableValue)'] = pd.to_numeric(df['BasicCostPrice(TaxableValue)'])
            df['IGSTRate'] = pd.to_numeric(df['IGSTRate'])
            df['IGSTAmt'] = pd.to_numeric(df['IGSTAmt'])
            df['Total Amount'] = pd.to_numeric(df['Total Amount'])
            df['Receiving Location'] = Receving_location
            df['PO Number'] = po_number
            df['PO Number'] = pd.to_numeric(df['PO Number'])

            # df = df[['POItem','ArticleEAN','Article Number','ArticleDescription','MRP','BasicCostPrice(TaxableValue)','Qty', 'UM','Total Amount','IGSTRate','IGSTAmt','Receiving Location','PO Number']]
            # Inserting extra cols 
            df.insert(loc = 4,column = 'HSNCode',value = '')
            # df.insert(loc = 8,column = 'UM',value = '')
            df.insert(loc = 9,column = 'TaxableValue',value = '')
            df.insert(loc = 11,column = 'CGSTRate',value = '')
            df.insert(loc = 12,column = 'CGSTAmt',value = '')
            df.insert(loc = 13,column = 'SGSTRate',value = '')
            df.insert(loc = 14,column = 'SGSTAmt',value = '')
            df.insert(loc = 15,column = 'UTGSTRate',value = '')
            df.insert(loc = 16,column = 'UTGSTAmt',value = '')
            df.insert(loc = 19,column = 'Vendor Penalty',value = '')
            df.insert(loc = 20,column = 'Vendor Name',value = '')
            # df.insert(loc = 23,column = 'Vendor Code',value = '')

            # Calculating Sum of Quantities
            Sum_of_Qty = df['Qty'].sum()
            
            if int(grand_total_text) != int(Sum_of_Qty):
                    print('Grand Total of the converted excel does not match with the pdf.')
                    logger.info('Grand Total of the converted excel does not match with the pdf.')
                    return
            else:
                # print("########## po_number: " + po_number)
                df.to_excel(base_path+"/99-Working/40-Extract-Excel/"+po_number+'.xlsx',index=False)
                print("Converted '" + file_name + "' to '"+po_number +".xlsx' in " + "{:.2f}".format(time.time() - startedProcessing, 2) + " seconds.")
                logger.info("Converted '" + file_name + "' to '"+po_number +".xlsx' in " + "{:.2f}".format(time.time() - startedProcessing, 2) + " seconds.")       
    
    
    except Exception as e:
        print('Error while po_exctract_data in '+ file_name + " " +str(e))
        logger.error('Error while po_exctract_data in '+ file_name + " " +str(e))
        exc_type, exc_obj, exc_tb = sys.exc_info()
        fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
        print(exc_type, fname, exc_tb.tb_lineno)

def getFilesToProcess_RRL(RootFolder, POSource, OrderDate, ClientCode, base_path):
    # converting str to datetime
    OrderDate = datetime.strptime(OrderDate, '%Y-%m-%d')
    # extracting year from the order date
    year = OrderDate.strftime("%Y")
    # formatting order date {2022-00-00) format
    OrderDate = OrderDate.strftime('%Y-%m-%d')
    try:
        inputFolderPath = base_path + "/99-Working/10-Download-Files/"
        outputFolderPath = base_path +"/99-Working/40-Extract-Excel/"
        startedProcessing = time.time()

        if len(os.listdir(inputFolderPath)) == 0:
            print("'"+inputFolderPath+"' Folder is empty, add pdf files to convert")
            logger.info("'"+inputFolderPath +
                        "' Folder is empty, add pdf files to convert")
            return
        else:
            logger.info("Converting PDF files to Excel...")
            print("Converting PDF files to Excel...")
            count = 0
            for f in os.listdir(inputFolderPath):
                if f.endswith('.pdf'):
                    with open(inputFolderPath+'/'+f, "rb") as pdf_file:
                        pdf_reader = PdfFileReader(pdf_file)
                        TotalPages = pdf_reader.numPages

                    # Remove last 7 pages
                    infile = PdfReader(inputFolderPath+'/'+f, 'rb')
                    output = PdfWriter()
                    for i in range(0,TotalPages-7):
                        p = infile.pages[i] 
                        output.add_page(p)
                    # Saving the new pdf after removing last pages
                    with open(base_path+'/99-Working/10-Download-Files/'+str(f), 'wb') as new_file:
                        output.write(new_file)
                        
                    # print(RootFolder, POSource, OrderDate, ClientCode, f, base_path)

                    pdfToTable_RRL(RootFolder, POSource, OrderDate, ClientCode, f, base_path)
                    count += 1
            print("Converted "+str(count)+" Files in " +
                    "{:.2f}".format(time.time() - startedProcessing, 2) + " seconds!")
            logger.info("Converted "+str(count)+" Files in " +
                            "{:.2f}".format(time.time() - startedProcessing, 2) + " seconds!")
                

    except Exception as e:
        print('Error while processing '+str(e))
        logger.error('Error while processing '+str(e))
        exc_type, exc_obj, exc_tb = sys.exc_info()
        fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
        print(exc_type, fname, exc_tb.tb_lineno)