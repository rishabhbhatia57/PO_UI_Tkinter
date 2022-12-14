import os
import shutil
from datetime import datetime
import pandas as pd
import xlsxwriter
import glob
import numpy as np
from openpyxl import load_workbook, Workbook
import openpyxl.utils.cell
import time
from config import ConfigFolderPath, CLIENTSFOLDERPATH
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import Alignment, Font
from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder
from openpyxl.utils import get_column_letter
# from openpyxl.styles import Style
import tabula
import csv
import json
import sys
from openpyxl.styles import PatternFill
from flask import jsonify

import BKE_log
from config import ITEMMASTERPATH, IGSTMASTERPATH, SGSTMASTERPATH, LOCATIONMASTERPATH, LOCATION2MASTERPATH, CLOSINGSTOCKMASTERPATH, get_client_code, get_client_name, REQSUMTEMPLATEPATH, TEMPLATESPATH, PACKINGSLIPTEMPLATEPATH
pd.options.mode.chained_assignment = None
# import warnings
# warnings.simplefilter(action='ignore', category=SettingWithCopyWarning)

logger = BKE_log.setup_custom_logger('root')

def getFilesToProcess(RootFolder, POSource, OrderDate, ClientCode, base_path):
    # converting str to datetime
    OrderDate = datetime.strptime(OrderDate, '%Y-%m-%d')
    # extracting year from the order date
    year = OrderDate.strftime("%Y")
    # formatting order date {2022-00-00) format
    OrderDate = OrderDate.strftime('%Y-%m-%d')
    try:
        inputFolderPath = base_path + "/99-Working/10-Download-Files/"
        outputFolderPath = base_path +"/99-Working/40-Extract-Excel/"
        startedProcessing = time.time()

        if len(os.listdir(inputFolderPath)) == 0:
            logger.info("'"+inputFolderPath +
                        "' Folder is empty, add pdf files to convert")
            print("'"+inputFolderPath+"' Folder is empty, add pdf files to convert")
            return
        else:
            logger.info("Converting PDF files to Excel...")
            print("Converting PDF files to Excel...")
            count = 0

            for f in os.listdir(inputFolderPath):
                if f.endswith('.pdf'):
                    fOutputExtension = f.replace('.pdf', '.xlsx')
                    pdfToTable(inputFolderPath+f, outputFolderPath+fOutputExtension,
                            RootFolder, POSource, OrderDate, ClientCode, f, base_path)
                    count += 1

            print("Converted "+str(count)+" Files in " +
                  "{:.2f}".format(time.time() - startedProcessing, 2) + " seconds!")
            logger.info("Converted "+str(count)+" Files in " +
                        "{:.2f}".format(time.time() - startedProcessing, 2) + " seconds!")
            # print("Completed in "+"{:.2f}".format(time.time() - startedProcessing,2)+ " seconds!")
    except Exception as e:
        logger.error("Error while processing files: "+str(e))
        exc_type, exc_obj, exc_tb = sys.exc_info()
        fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
        print(exc_type, fname, exc_tb.tb_lineno)
        print("Error while processing files: "+str(e))


def pdfToTable(inputPath, outputPath, RootFolder, POSource, OrderDate, ClientCode, filecsv, base_path):

    try:
        # logger.info("Converting '"+ filecsv)# +"' to excel '"+filecsv.replace('.pdf', '.xlsx')+"'")
        # print("Converting '"+ filecsv)# +"' to excel '"+filecsv.replace('.pdf', '.xlsx')+"'")
        # converting str to datetimedatetime
        OrderDate = datetime.strptime(OrderDate, '%Y-%m-%d')
        # extracting year from the order date
        year = OrderDate.strftime("%Y")
        # formatting order date {2022-00-00) format
        OrderDate = OrderDate.strftime('%Y-%m-%d')

        startedProcessing = time.time()

        intermediateCSV = base_path +"/99-Working/30-Extract-CSV/"+filecsv.replace('.pdf', '.csv')
        intermediateExcel = base_path +"/99-Working/20-Intermediate-Files/1_" +filecsv.replace('.pdf', '.xlsx')
        intermediateExcel2 = base_path +"/99-Working/20-Intermediate-Files/2_" +filecsv.replace('.pdf', '.xlsx')
        intermediateoutputPath = base_path +"/99-Working/20-Intermediate-Files/3_" +filecsv.replace('.pdf', '.xlsx')

        try:
            tabula.convert_into(input_path=inputPath, output_path=intermediateCSV, pages='all', lattice=True)
        except Exception as e:
            print("Error while converting file "+str(e))
            logger("Error while converting file "+inputPath+" "+str(e))
            return
        # making a dataframe using csv
        df = pd.read_csv(filepath_or_buffer=intermediateCSV,skiprows=[1, 2, 3])

        # converting csv to excel
        writer = pd.ExcelWriter(intermediateExcel, engine='xlsxwriter')
        df.to_excel(writer, sheet_name='Sheet1', merge_cells=False)
        writer.save()

        # loading excel
        workbook = load_workbook(filename=intermediateExcel)
        sheet = workbook.active

        # finding the last row needed
        endrow = 1
        endcolumn = 1

        for i in range(1, 10000000):
            if sheet.cell(i, 1).value != None and "Grand" in sheet.cell(i, 1).value:
                endrow = i
                break

        for i in range(1, 10000000):
            if sheet.cell(2, i).value != None and "Total" in sheet.cell(2, i).value:
                endcolumn = i
                break

        # print(endrow, endcolumn)

        # Cleaning loop
        for i in range(1, endrow+2):
            for j in range(1, endcolumn+1):
                if sheet.cell(i, j).value != None and "Aditya" in sheet.cell(i, j).value:
                    sheet.cell(i, 1).value = "remove"
                elif sheet.cell(i, j).value != None and "P. O. Number" in sheet.cell(i, j).value:
                    sheet.cell(i, 1).value = "remove"
                elif sheet.cell(i, j).value != None and "PO" in sheet.cell(i, j).value and i != 2:
                    sheet.cell(i, 1).value = "remove"
                elif sheet.cell(i, j).value != None and "_x000D_" in sheet.cell(i, j).value:
                    temp = sheet.cell(i, j).value.replace("_x000D_", "")
                    sheet.cell(i, j).value = temp

        # rearranging the data into a single row

        # column numbers
        CGSTRate = endcolumn+1
        CGSTAmt = endcolumn+2
        SGSTRate = endcolumn+3
        SGSTAmt = endcolumn+4
        UTGSTRate = endcolumn+5
        UTGSTAmt = endcolumn+6
        IGSTRate = endcolumn+7
        IGSTAmt = endcolumn+8
        VendorPenalty = endcolumn+9
        VendorName = endcolumn+8
        ReceivingLocation = endcolumn+9
        PONumber = endcolumn+10
        VendorCode = endcolumn+11
        for i in range(1, endrow):
            for j in range(1, endcolumn+1):
                if sheet.cell(i, j).value == "CGST":
                    sheet.cell(i-1, CGSTRate).value = sheet.cell(i, j+1).value
                    sheet.cell(i-1, CGSTAmt).value = sheet.cell(i, j+2).value
                    sheet.cell(i, j+1).value = ""
                    sheet.cell(i, j+2).value = ""
                    sheet.cell(i, j).value = ""
                elif sheet.cell(i, j).value == "SGST":
                    sheet.cell(i-2, SGSTRate).value = sheet.cell(i, j+1).value
                    sheet.cell(i-2, SGSTAmt).value = sheet.cell(i, j+2).value
                    sheet.cell(i, j+1).value = ""
                    sheet.cell(i, j+2).value = ""
                    sheet.cell(i, j).value = ""
                elif sheet.cell(i, j).value == "UTGST":
                    sheet.cell(i-3, UTGSTRate).value = sheet.cell(i, j+1).value
                    sheet.cell(i-3, UTGSTAmt).value = sheet.cell(i, j+2).value
                    sheet.cell(i, j+1).value = ""
                    sheet.cell(i, j+2).value = ""
                    sheet.cell(i, j).value = ""
                elif sheet.cell(i, j).value == "IGST":
                    sheet.cell(i-4, IGSTRate).value = sheet.cell(i, j+1).value
                    sheet.cell(i-4, IGSTAmt).value = sheet.cell(i, j+2).value
                    sheet.cell(i, j+1).value = ""
                    sheet.cell(i, j+2).value = ""
                    sheet.cell(i, j).value = ""
                elif sheet.cell(i, j).value == "VendorPenalty":
                    sheet.cell(
                        i-4, VendorPenalty).value = sheet.cell(i, j+1).value
                    sheet.cell(i, j+1).value = ""
                    sheet.cell(i, j).value = ""

        # removing empty and unwanted rows
        counter = endrow
        while counter > 0:
            if sheet.cell(counter, 1).value == None or sheet.cell(counter, 1).value == "" or sheet.cell(counter, 1).value == "remove":
                sheet.delete_rows(counter)
            counter = counter - 1

        sheet.delete_cols(endcolumn-1)
        sheet.delete_cols(endcolumn-2)

        sheet["A1"] = "POItem"
        sheet["B1"] = "ArticleEAN"
        sheet["C1"] = "Article Number"
        sheet["D1"] = "ArticleDescription"
        sheet["E1"] = "HSNCode"
        sheet["F1"] = "MRP"
        sheet["G1"] = "BasicCostPrice(TaxableValue)"
        sheet["H1"] = "Qty"
        sheet["I1"] = "UM"
        sheet["J1"] = "TaxableValue"
        # sheet["K1"] = "GSTRate"
        # sheet["L1"] = "GSTAmt"
        sheet["K1"] = "Total Amount"
        sheet["L1"] = "CGSTRate"
        sheet["M1"] = "CGSTAmt"
        sheet["N1"] = "SGSTRate"
        sheet["O1"] = "SGSTAmt"
        sheet["P1"] = "UTGSTRate"
        sheet["Q1"] = "UTGSTAmt"
        sheet["R1"] = "IGSTRate"
        sheet["S1"] = "IGSTAmt"
        sheet["T1"] = "Vendor Penalty"
        sheet["U1"] = "Vendor Name"
        sheet["V1"] = "Receiving Location"
        sheet["W1"] = "PO Number"
        sheet["X1"] = "Vendor Code"

        # workbook = xlsxwriter.Workbook(intermediateExcel, {'strings_to_numbers': True})
        df2 = pd.read_csv(filepath_or_buffer=intermediateCSV,
                          on_bad_lines='skip')

        # print(df2)
        writer2 = pd.ExcelWriter(intermediateExcel2, engine='xlsxwriter')
        df2.to_excel(writer2, sheet_name='Sheet1', merge_cells=False)
        writer2.save()
        workbook2 = load_workbook(filename=intermediateExcel2)
        sheet2 = workbook2.active

        # Cleaning loop
        for i in range(1, 10):
            for j in range(1, 5):
                if sheet2.cell(i, j).value != None and "_x000D_" in sheet2.cell(i, j).value:
                    temp = sheet2.cell(i, j).value.replace("_x000D_", "%")
                    sheet2.cell(i, j).value = temp

        # Recalculating endrow
        for i in range(1, 10000000):
            if sheet.cell(i, 1).value != None and "Grand" in sheet.cell(i, 1).value:
                endrow = i
                break

        PONumberRow = 1
        for i in range(1, 25):
            if sheet2.cell(i, 1).value != None and "P. O. Number" in sheet2.cell(i, 1).value:
                PONumberRow = i
                break

        VendorNameValue = sheet2.cell(
            2, 1).value.replace(":", "").split('%')[1]
        VendorCodeValue = sheet2.cell(
            2, 1).value.replace(":", "").split('%')[3]
        ReceivingLocationValue = sheet2.cell(
            4, 1).value.replace(":", "").split('%')[3]
        PONumValue = sheet2.cell(PONumberRow, 1).value.replace(
            ":", "").split('%')[1]
        for i in range(2, endrow):
            sheet.cell(i, VendorName).value = VendorNameValue
            sheet.cell(i, ReceivingLocation).value = ReceivingLocationValue
            sheet.cell(i, PONumber).value = PONumValue
            sheet.cell(i, VendorCode).value = VendorCodeValue

        # get  Delete row count
        lastrow = 1
        for i in range(endrow, 1000000):
            if sheet.cell(i, 1).value != None and "Other Conditions" in sheet.cell(i, 1).value:
                lastrow = i+5
                break

        while endrow <= lastrow:
            sheet.delete_rows(lastrow)
            lastrow = lastrow - 1

        workbook2.save(filename=intermediateoutputPath)
        workbook.save(filename=outputPath)

        # logger.info("Converted '"+ inputPath + "' to '" + outputPath+"'"+ " in "+ "{:.2f}".format(time.time() - startedProcessing,2)+ " seconds.")
        # print("Converted '"+ inputPath + "' to '" + outputPath+"'"+ " in "+ "{:.2f}".format(time.time() - startedProcessing,2)+ " seconds.")
        print("Converted '" + filecsv + "' to '"+filecsv.replace('pdf', 'xlsx') +
              "' in " + "{:.2f}".format(time.time() - startedProcessing, 2) + " seconds.")
        logger.info("Converted '" + filecsv + "' to '"+filecsv.replace('pdf', 'xlsx') +
                    "' in " + "{:.2f}".format(time.time() - startedProcessing, 2) + " seconds.")
        return "Conversion Complete!"

    except Exception as e:
        logger.error("Error while processing file: "+str(e))
        exc_type, exc_obj, exc_tb = sys.exc_info()
        fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
        print(exc_type, fname, exc_tb.tb_lineno)
        print("Error while processing file: "+str(e))