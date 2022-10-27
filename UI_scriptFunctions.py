import tkinter as tk                    
from tkinter import ttk
from tkinter import *
from screeninfo import get_monitors
from tkcalendar import DateEntry
from tkinter import filedialog as fd
from tkinter.messagebox import showinfo
import tkinter.scrolledtext as st
import os
import threading
import base64
import tkinter.font as tkFont
import json
from openpyxl import load_workbook,Workbook
import openpyxl.utils.cell
import pyperclip
from datetime import datetime
import sys
from config import ConfigFolderPath,CLIENTSFOLDERPATH, headingFont,fieldFont,buttonFont,labelFont,pathFont,logFont, PKG_CLIENTS
from BKE_mainFunction import startProcessing
import threading
import BKE_log
import datetime

logger = BKE_log.setup_custom_logger('root')



def select_folder(showPath):
    try:
        filetypes = (
            ('pdf files', '*.pdf'),
            ('All files', '*.*')
        )
        global POFolderSelected 
        POFolderSelected = fd.askdirectory()    
        if POFolderSelected == '':
            showPath.config(text='No Folder selected')
            showinfo(
                title='Invalid Selection',
                message='Please select the folder.'
            )
        else:
            showPath.config(text=POFolderSelected)
            showinfo(
                title='Selected Folder',
                message=POFolderSelected
            )
            return POFolderSelected

    except Exception as e:
        print(e)
        exc_type, exc_obj, exc_tb = sys.exc_info()
        fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
        print(exc_type, fname, exc_tb.tb_lineno)
        logger.error('Error while selection: '+str(e))

def select_files(showPath,showReqPath,showOrderdate):
    try:

        filetypes = (
            ('Excel Workbook', '*.xlsx'),
            ('All files', '*.*')
        )
        global ReqFileSelected 
        ReqFileSelected = fd.askopenfilename()
        if not ReqFileSelected.lower().endswith('.xlsx'):
            showinfo(title='Invalid Selection',
            message='Wrong file selected. Only excel workbook with extenstion ".xlsx" can be selected.')
        else:
            showinfo(title='Please wait',message="Fetching Client Name and Order Date...\nClick OK to continue.")
            showPath.config(text=ReqFileSelected)
            ReqSumWorkbook = load_workbook(ReqFileSelected,data_only=True)
            ReqSumSheet = ReqSumWorkbook.active
            print(PKG_CLIENTS,ReqSumSheet.cell(1,2).value,ReqSumSheet.cell(2,2).value)

            # checks if client name exists in client.sjon or not
            if ReqSumSheet.cell(1,2).value not in PKG_CLIENTS:
                logger.error('Could not find the client name in client.Json file. Please check requirement summary before proceeding.')
            else:
                showReqPath.set(ReqSumSheet.cell(1,2).value)

            # checks if Date is correct/ or in cirrect format
            try:
                datetime.datetime.strptime(str(ReqSumSheet.cell(2,2).value), '%Y-%m-%d')
                showOrderdate.set(ReqSumSheet.cell(2,2).value)
            except ValueError:
                logger.error('Could not find the order date in requirement summary file. Accepted Date format is YYYY-MM-DD. Please check requirement summary before proceeding.')
            
        return ReqFileSelected
    except Exception as e:
        print(e)
        exc_type, exc_obj, exc_tb = sys.exc_info()
        fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
        print(exc_type, fname, exc_tb.tb_lineno)
        logger.error('Error while selection: '+str(e))

def open_folder(params,frame):
    try: 
    
        year = params[2].strftime('%Y')
        date = str(params[2])
        

        path = params[0]+'/'+params[1]+'-'+year+'/'+date+'/'+params[3]
        isExist = os.path.exists(path)
        if not isExist:
            showinfo(
                title='Invalid Selection',
                message="Folder doesn't exists. Check Client Name, path or Order Date Selected"
            )
            print("Path doesn't exists. Please check date or client name.")
        else:
            pyperclip.copy(path)
            # RequirementSummaryPath = Label(frame,text=path,wraplength=800)
            # RequirementSummaryPath.grid(row=5,column=2,padx=20, pady=20,sticky=W)
            showinfo(title='Path Copied',message="'"+path+"' is copied to the clipboard.")
            # path = os.path.realpath(path)
            # os.startfile(path)
    except Exception as e:
        print(e)
        exc_type, exc_obj, exc_tb = sys.exc_info()
        fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
        print(exc_type, fname, exc_tb.tb_lineno)
        logger.error('Error while opening folder: '+str(e))

def open_folder_packaging(params,frame):
    print(params)
    try:
            
        if  params[2] == '' or params[3] == '':
            showinfo(
                title='Invalid Selection',
                message="No folder is selected. Check Client Name, path or Order Date Selected"
            )
        else:

            year = datetime.datetime.strptime(params[2], '%Y-%m-%d').strftime('%Y')
            date = str(params[2])
            

            path = params[0]+'/'+params[1]+'-'+year+'/'+date+'/'+params[3]
            isExist = os.path.exists(path)
            if not isExist:
                showinfo(
                    title='Invalid Selection',
                    message="Folder doesn't exists. Check Client Name, path or Order Date Selected"
                )
                print("Path doesn't exists. Please check date or client name.")
            else:
                pyperclip.copy(path)
                showinfo(title='Path Copied',message="'"+path+"' is copied to the clipboard.")
                # path = os.path.realpath(path)
                # os.startfile(path)

    except Exception as e:
        print(e)
        exc_type, exc_obj, exc_tb = sys.exc_info()
        fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
        print(exc_type, fname, exc_tb.tb_lineno)
        logger.error('Error while opening folder: '+str(e))
    
def begin_order_processing(mode, client, date, path, consoleLabel, thread_name):

    try:
        print(mode, client, date, path, thread_name)
        
        # with open(CLIENTSFOLDERPATH, 'r') as jsonFile:
        #         config = json.load(jsonFile)
        #         ClientCode = config

        ClientCodeSelected = client
        OrderDateSelected = date
        requestedpath = path
        if OrderDateSelected == '' or requestedpath == 'No Folder selected':
            showinfo(
            title='Invalid Selection',
            message="Invalid Client Name, path or Order Date Selected"
        )
        else:
            with open(ConfigFolderPath, 'r') as jsonFile:
                config = json.load(jsonFile)    
                # pythonenvpath = config['pythonPath']
                # pythonScriptPath = config['appsScriptPath']

            if ClientCodeSelected == '--select--':
                print('Wrong input')
                showinfo(
                    title='Invalid Selection',
                    message="Please selected Client Name from dropdowm menu."
                )
            else:
                consoleLabel.config(text='Console logs    -    🔄    -    Processing...')
                startProcessing(mode=mode,clientname=ClientCodeSelected,orderdate=str(OrderDateSelected),processing_source=requestedpath)
                consoleLabel.config(text='Console logs    -    ✅    -    Completed!')
    except Exception as e:
        print(e)
        exc_type, exc_obj, exc_tb = sys.exc_info()
        fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
        print(exc_type, fname, exc_tb.tb_lineno)
        logger.error('Error while processing: '+str(e))