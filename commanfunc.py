import tkinter as tk                    
from tkinter import ttk
from tkinter import *
from screeninfo import get_monitors
from tkcalendar import DateEntry
from tkinter import filedialog as fd
from tkinter.messagebox import showinfo
import tkinter.scrolledtext as st
import os
import threading
import base64
import tkinter.font as tkFont
import json
from openpyxl import load_workbook,Workbook
import openpyxl.utils.cell
import pyperclip

path1 = 'C:/Users/HP/Desktop/PO Metadata/Configfiles-Folder/'

def select_folder(showPath):
    filetypes = (
        ('pdf files', '*.pdf'),
        ('All files', '*.*')
    )
    global POFolderSelected 
    POFolderSelected = fd.askdirectory()
    showPath.config(text=POFolderSelected)
    
    showinfo(
        title='Selected Folder',
        message=POFolderSelected
    )
    return POFolderSelected

def select_files(showPath,showReqPath,showOrderdate):
    filetypes = (
        ('Excel Workbook', '*.xlsx'),
        ('All files', '*.*')
    )
    global ReqFileSelected 
    ReqFileSelected = fd.askopenfilename()
    if not ReqFileSelected.lower().endswith('.xlsx'):
        showinfo(title='Invalid Selection',
        message='Wrong file selected. Only excel workbook with extenstion ".xlsx" can be selected.')
    else:
        showinfo(title='Please wait',message="Fetching Client Name and Order Date...")
        showPath.config(text=ReqFileSelected)

        ReqSumWorkbook = load_workbook(ReqFileSelected,data_only=True) 
        ReqSumSheet = ReqSumWorkbook.active
        showReqPath.config(text=ReqSumSheet.cell(2,6).value)
        showOrderdate.config(text=ReqSumSheet.cell(2,4).value)
    
    return ReqFileSelected

def openfolder(params,frame):
    
    year = params[2].strftime('%Y')
    date = str(params[2])
    

    path = params[0]+'/'+params[1]+'-'+year+'/'+date+'/'+params[3]
    isExist = os.path.exists(path)
    if not isExist:
        showinfo(
            title='Invalid Selection',
            message="Folder doesn't exists. Check Client Name, path or Order Date Selected"
        )
        print("Path doesn't exists. Please check date or client name.")
    else:
        pyperclip.copy(path)
        RequirementSummaryPath = Label(frame,text=path)
        RequirementSummaryPath.grid(row=5,column=2,padx=20, pady=20,sticky=W)
        showinfo(title='Path Copied',message="'"+path+"' is copied to the clipboard.")
        # path = os.path.realpath(path)
        # os.startfile(path)
    

def selectedFun(mode, client, date,path):
    print(path)
    with open(path1+'/'+'client.json', 'r') as jsonFile:
            config = json.load(jsonFile)
            print(config)
            ClientCode = config

    ClientCodeSelected = client
    OrderDateSelected = date
    requestedpath = path
    if ClientCodeSelected == '' or OrderDateSelected == '' or requestedpath == '':
        showinfo(
        title='Invalid Selection',
        message="Invalid Client Name, path or Order Date Selected"
    )
    else:
        with open(path1+'config.json', 'r') as jsonFile:
            config = json.load(jsonFile)    
            pythonenvpath = config['pythonPath']
            pythonScriptPath = config['appsScriptPath']

        encodedClientCodeSelected = ClientCode[ClientCodeSelected]
        encodedOrderDateSelected = str(OrderDateSelected).replace(' ', "#")
        enodedPOFolderSelected = requestedpath.replace(' ', "#") 

        # Running code on CMD
        print(pythonenvpath,pythonScriptPath,mode,encodedClientCodeSelected,encodedOrderDateSelected,enodedPOFolderSelected)
        os.system(pythonenvpath +" "+ pythonScriptPath+" "+mode+" "+encodedClientCodeSelected+" "+encodedOrderDateSelected+" "+enodedPOFolderSelected)


        